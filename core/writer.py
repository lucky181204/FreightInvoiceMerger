"""Writer — writes extracted invoice data into the template Excel file.

Only modifies cell.value. Preserves all formatting, fonts, borders, formulas,
column widths, print settings, freeze panes, filters, merged cells.
"""

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from utils.logger import logger
from rules.registry import RuleRegistry


def write_to_template(
    template_path: str,
    output_path: str,
    data: list[dict],
    rule_id: str,
    auto_sort: bool = True,
):
    """
    Write extracted invoice data into the template starting at the rule's start_row.
    Sorts rows in-place on the sheet by the specified sort column.
    Only modifies cell values — all formatting is preserved.
    """
    rule = RuleRegistry.get_rule(rule_id)
    if not rule:
        raise ValueError(f"Rule not found: {rule_id}")

    start_row = rule.RULE_META.get("start_row", 5)
    sort_col_letter = rule.RULE_META.get("sort_column", "U")

    # Load template (writable mode, preserve everything)
    wb = load_workbook(template_path)
    ws = wb.active
    logger.info(f"写入模板，起始行：{start_row}")

    # ── Step 1: Write all data to sheet ──
    sort_col_idx = column_index_from_string(sort_col_letter)
    for i, row_data in enumerate(data):
        row_num = start_row + i
        for col_letter, value in row_data.items():
            try:
                col_idx = column_index_from_string(col_letter)
                ws.cell(row=row_num, column=col_idx).value = value
            except Exception as e:
                logger.warning(f"写入失败：列{col_letter} 行{row_num} — {e}")
        logger.info(f"写入第{row_num}行")

    # ── Step 2: Sort rows in-place if enabled ──
    if auto_sort and data:
        logger.info(f"按{sort_col_letter}列排序...")
        _sort_sheet_rows(ws, start_row, start_row + len(data) - 1, sort_col_idx)

    # ── Step 3: Save ──
    wb.save(output_path)
    wb.close()
    logger.info(f"保存完成：{output_path}")
    return output_path


def _sort_sheet_rows(ws, first_row: int, last_row: int, sort_col_idx: int):
    """
    Sort rows in the worksheet in-place.
    Reads all cell values in the range, sorts by sort column, writes back.
    Preserves all formatting because we only modify cell.value.
    """
    if first_row >= last_row:
        return

    # Determine the max column in use across the sort range
    max_col = ws.max_column
    # Also scan the range for meaningful column count
    for r in range(first_row, last_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                max_col = max(max_col, c)

    # Read all rows into list of tuples
    rows = []
    for r in range(first_row, last_row + 1):
        row_vals = []
        sort_val = None
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
            if c == sort_col_idx:
                sort_val = v
        rows.append((sort_val, row_vals))

    # Sort by sort value (numeric if possible, else text)
    def sort_key(item):
        val = item[0]
        try:
            return (0, float(val))
        except (ValueError, TypeError):
            return (1, str(val or ""))

    rows.sort(key=sort_key)

    # Write back sorted values
    for i, (_, row_vals) in enumerate(rows):
        r = first_row + i
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = row_vals[c - 1]

    logger.info(f"排序完成 ({first_row}-{last_row}行)")
