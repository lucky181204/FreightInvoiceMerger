"""Trade terms manifest parser.

Matches BL No (提单号) in the trade terms manifest to find:
- Customer PO → C10
- Business No → H7
- Scheduled sailing → H5, H6
"""

import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def find_by_blno(manifest_path: str, bl_no: str) -> dict | None:
    """
    Search trade terms manifest for the given BL number.

    Manifest columns:
      A: 客户PO编号
      F: 提单号
      J: 预定船期
      L: 业务编号

    Returns dict with keys: po, business_no, sailing_date
    or None if not found.
    """
    if not bl_no:
        return None

    wb = load_workbook(manifest_path, data_only=True)
    try:
        ws = wb.active

        # Locate column positions
        po_col = None
        bl_col = None
        sail_col = None
        biz_col = None

        for c in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(row=1, column=c).value
            if v and isinstance(v, str):
                vc = v.strip().replace(" ", "").replace("　", "")
                if "客户PO" in vc or "PO编号" in vc:
                    po_col = c
                elif "提单号" in vc or "BL" in vc.upper():
                    bl_col = c
                elif "预定船期" in vc or "船期" in vc:
                    sail_col = c
                elif "业务编号" in vc or "业务号" in vc:
                    biz_col = c

        if not bl_col:
            # Try alternate column names
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row=1, column=c).value
                if v and isinstance(v, str) and ("提单" in v):
                    bl_col = c
                elif v and isinstance(v, str) and ("PO" in v.replace(" ", "")):
                    po_col = c

        if not bl_col:
            logger.warning("贸易条款清单未找到提单号列")
            wb.close()
            return None

        # Search for BL No in data rows
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=bl_col).value
            if cell_val and str(bl_no) in str(cell_val):
                result = {}
                if po_col:
                    pv = ws.cell(row=r, column=po_col).value
                    result["po"] = str(pv).strip() if pv else ""
                if biz_col:
                    bv = ws.cell(row=r, column=biz_col).value
                    result["business_no"] = str(bv).strip() if bv else ""
                if sail_col:
                    sv = ws.cell(row=r, column=sail_col).value
                    if sv:
                        # Handle datetime object
                        if hasattr(sv, 'strftime'):
                            result["sailing_date"] = sv.strftime("%Y/%m/%d")
                        else:
                            result["sailing_date"] = str(sv).strip()
                    else:
                        result["sailing_date"] = ""
                wb.close()
                return result

        wb.close()
        return None
    except Exception as e:
        logger.warning(f"贸易条款清单读取失败: {e}")
        try:
            wb.close()
        except Exception:
            pass
        return None
