"""Manifest (舱单/container data) parser.

Dynamically finds container data by scanning ALL worksheets for
container number (箱号/柜号) and container type (箱型) headers.

Supports:
- Any workbook name
- Any sheet name
- Any header position
- BL No filtering when available
- Numeric container type combining
"""

import logging
import re
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

EMPTY_ROW_LIMIT = 5


def normalize_header(value) -> str:
    """Normalize header text for matching."""
    text = str(value or "")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def is_container_no_header(value) -> bool:
    """Check if header text indicates a container number column."""
    text = normalize_header(value)
    if "箱号" in text or "柜号" in text:
        return True
    if "container no" in text or "container number" in text:
        return True
    if "集装箱号" in text:
        return True
    return False


def is_container_type_header(value) -> bool:
    """Check if header text indicates a container type column."""
    text = normalize_header(value)
    if "箱型" in text:
        return True
    if "container type" in text or "container size" in text:
        return True
    return False


def is_bl_no_header(value) -> bool:
    """Check if header text indicates a BL number column."""
    text = normalize_header(value)
    if "提单号" in text or "主提单号" in text:
        return True
    if "hbl no" in text or "b/l no" in text or "bl no" in text:
        return True
    if "booking no" in text:
        return True
    return False


def find_containers_by_blno(manifest_path: str, bl_no: str, progress_callback=None) -> list[dict]:
    """
    Search manifest file for container data matching BL No.

    Scans ALL worksheets, finds container number and type columns dynamically.

    Returns list of dicts with:
      container_no (str), container_type (str)
    """
    if not bl_no:
        return []

    def log(msg):
        logger.info(msg)
        if progress_callback:
            progress_callback(msg)

    log(f"读取舱单文件: {manifest_path}")

    wb = load_workbook(manifest_path, data_only=True)
    candidates = []

    # Step 1: Scan all worksheets for header positions
    for ws in wb.worksheets:
        sheet_name = ws.title
        scan_info = _scan_sheet(ws)
        if scan_info:
            scan_info["sheet_name"] = sheet_name
            candidates.append(scan_info)

    if not candidates:
        log("未在上传的Excel文件中找到箱号或柜号列，请检查文件模板。")
        wb.close()
        return []

    # Step 2: Select best candidate sheet
    best = _select_best_sheet(candidates)

    if len(candidates) > 1:
        log(f"找到多个候选工作表，已选择: {best['sheet_name']}")

    ws = wb[best["sheet_name"]]
    cntr_col = best["cntr_col"]
    type_col = best.get("type_col")
    bl_col = best.get("bl_col")
    data_start = best["data_start"]

    log(f"数据工作表: {best['sheet_name']}")
    log(f"箱号标题位置: {_col_letter(cntr_col)}{best['cntr_row']}")
    if type_col:
        log(f"箱型标题位置: {_col_letter(type_col)}{best['type_row']}")
    if bl_col:
        log(f"提单号列: {_col_letter(bl_col)}")

    # Step 3: Read data rows
    results = []
    empty_count = 0
    in_matching_bl = not bl_col  # If no BL column, match all rows
    current_bl = None
    max_row = ws.max_row

    for r in range(data_start, max_row + 1):
        # Read cell values
        cntr_val = ws.cell(row=r, column=cntr_col).value

        # Check for end conditions
        if cntr_val is None or str(cntr_val).strip() == "":
            empty_count += 1
            if empty_count >= EMPTY_ROW_LIMIT:
                break
            continue
        empty_count = 0

        # Check for total/summary row
        cntr_str = str(cntr_val).strip()
        if cntr_str.upper() in ("TOTAL", "合计", "总计"):
            break

        # BL No filtering
        if bl_col:
            bl_val = ws.cell(row=r, column=bl_col).value
            if bl_val and str(bl_val).strip():
                current_bl = str(bl_val).strip()
                in_matching_bl = (current_bl == bl_no)

        if not in_matching_bl:
            continue

        # Container type
        type_val = ""
        if type_col:
            tv = ws.cell(row=r, column=type_col).value
            next_tv = ws.cell(row=r, column=type_col + 1).value if type_col + 1 <= max_row else None
            type_val = _build_container_type(tv, next_tv)

        results.append({
            "container_no": cntr_str,
            "container_type": type_val,
        })

    wb.close()
    logger.info(f"提取{len(results)}个集装箱数据")
    return results


def _scan_sheet(ws) -> dict | None:
    """Scan a worksheet for container number and type headers.
    Returns dict with positions or None if not found."""
    max_row = ws.max_row
    max_col = ws.max_column

    # Find header candidates
    headers = []  # list of (row, col, type)

    for r in range(1, min(max_row + 1, 200)):
        for c in range(1, min(max_col + 1, 100)):
            v = ws.cell(row=r, column=c).value
            if v is None or not isinstance(v, str):
                continue

            if is_container_no_header(v):
                headers.append({"row": r, "col": c, "type": "cntr"})
            if is_container_type_header(v):
                headers.append({"row": r, "col": c, "type": "type"})
            if is_bl_no_header(v):
                headers.append({"row": r, "col": c, "type": "bl"})

    cntr_headers = [h for h in headers if h["type"] == "cntr"]
    type_headers = [h for h in headers if h["type"] == "type"]
    bl_headers = [h for h in headers if h["type"] == "bl"]

    if not cntr_headers:
        return None

    # Try to find container + type headers close together
    best_cntr = cntr_headers[0]
    best_type = None
    best_bl = bl_headers[0] if bl_headers else None

    # Look for type header within ±2 rows of cntr header
    for th in type_headers:
        if abs(th["row"] - best_cntr["row"]) <= 2:
            best_type = th
            break

    if not best_type and type_headers:
        # Just use the first type header found
        best_type = type_headers[0]

    data_start = best_cntr["row"] + 1
    if best_type:
        data_start = max(data_start, best_type["row"] + 1)

    return {
        "sheet_name": ws.title,
        "cntr_col": best_cntr["col"],
        "cntr_row": best_cntr["row"],
        "type_col": best_type["col"] if best_type else None,
        "type_row": best_type["row"] if best_type else None,
        "bl_col": best_bl["col"] if best_bl else None,
        "bl_row": best_bl["row"] if best_bl else None,
        "data_start": data_start,
        "cntr_count": len(cntr_headers),
        "type_count": len(type_headers),
    }


def _select_best_sheet(candidates: list) -> dict:
    """Select the best candidate sheet from multiple options."""
    # Prefer sheets with both cntr and type columns
    complete = [c for c in candidates if c["type_col"]]
    if complete:
        candidates = complete

    # Prefer sheets with more data rows
    candidates.sort(key=lambda c: c.get("data_start", 0), reverse=False)
    # Between same data_start, prefer those with BL column
    candidates.sort(key=lambda c: 1 if c.get("bl_col") else 0, reverse=True)

    return candidates[0]


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to letter(s)."""
    result = ""
    while col_idx > 0:
        col_idx -= 1
        result = chr(ord("A") + col_idx % 26) + result
        col_idx //= 26
    return result


def _normalize_excel_value(value) -> str:
    """Normalize a cell value for processing."""
    if value is None:
        return ""
    if isinstance(value, float) and value == value and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            pass
    return text


def _build_container_type(type_value, next_column_value=None) -> str:
    """
    Build container type string.
    - If pure digits, combine with next column value
    - Otherwise use as-is
    """
    raw_type = _normalize_excel_value(type_value)
    raw_next = _normalize_excel_value(next_column_value)

    if not raw_type:
        return ""

    if re.fullmatch(r"\d+", raw_type):
        return f"{raw_type}{raw_next}"

    return raw_type
