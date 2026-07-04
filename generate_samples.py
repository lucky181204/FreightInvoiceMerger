"""Generate a sample template Excel and sample invoice ZIP for testing."""

import os
import zipfile
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Directories ──
OUTPUT = Path(__file__).parent

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill('solid', fgColor='E6F4FF')
HEADER_FONT = Font(name='Microsoft YaHei UI', bold=True, size=10)
DATA_FONT = Font(name='Microsoft YaHei UI', size=10)


def create_template():
    """Create the template file: Freight Invoice list 2026Fareast.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Freight Invoice list 2026Fareast'

    headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
               'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']

    # Row 4: header row
    header_texts = {
        'A': '序号', 'D': 'Invoice No.', 'E': 'HBL No.', 'F': 'MBL No.',
        'I': 'Volume(CBM)', 'K': '起运港', 'L': '目的港',
        'M': 'Vessel/Voyage', 'O': 'ETD', 'P': 'Container No.',
        'Q': 'Invoice No.', 'S': 'Gross Weight(KGS)', 'T': 'Description',
        'U': 'HBL No.',
    }

    for col_letter, text in header_texts.items():
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=4, column=col_idx)
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Set column widths
    col_widths = {'A': 6, 'B': 12, 'C': 16, 'D': 22, 'E': 22, 'F': 22,
                  'G': 12, 'H': 14, 'I': 14, 'J': 12, 'K': 16, 'L': 16,
                  'M': 24, 'N': 14, 'O': 14, 'P': 24, 'Q': 22, 'R': 12,
                  'S': 18, 'T': 40, 'U': 22}

    for col_letter, w in col_widths.items():
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        ws.column_dimensions[col_letter].width = w

    # Freeze pane at row 5
    ws.freeze_panes = 'A5'

    path = OUTPUT / 'resources' / 'Freight Invoice list 2026Fareast.xlsx'
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f'Template created: {path}')
    return str(path)


def create_sample_invoice(invoice_num: str, hbl: str, mbl: str, volume: str,
                          pol: str, pod: str, vessel: str, etd: str,
                          container: str, gross_weight: str,
                          description: str) -> Workbook:
    """Create a single sample invoice workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'

    # ── Layout matching PRD fields ──
    ws['A1'] = 'FREIGHT INVOICE'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['B5'] = f'Shipper: {pol},'
    ws['B6'] = f'Consignee: {pod},'
    ws['B10'] = f'Vessel: {vessel}'
    ws['B12'] = f'HBL: {hbl}'
    ws['B5'].font = DATA_FONT
    ws['B6'].font = DATA_FONT
    ws['B10'].font = DATA_FONT
    ws['B12'].font = DATA_FONT

    ws['C10'] = container
    ws['D10'] = gross_weight
    ws['E10'] = f'{volume}V'

    ws['F10'] = 'AAA'
    ws['F11'] = 'BBB'
    ws['F12'] = 'CCC'
    ws['F13'] = 'DDD'

    ws['H6'] = f'HBL: {hbl}'
    ws['H7'] = f'MBL: {mbl}'
    ws['I29'] = invoice_num

    for row in ws.iter_rows(min_row=1, max_row=30, max_col=8):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = DATA_FONT

    return wb


def create_sample_zip():
    """Create a ZIP archive with sample invoice files."""
    zip_dir = OUTPUT / 'resources'
    zip_dir.mkdir(parents=True, exist_ok=True)
    zip_path = zip_dir / 'sample_invoices.zip'

    invoice_data = [
        ('INV-2026-001', 'HBL001', 'MBL-ABC', '25.5', 'Shanghai',
         'Los Angeles', 'MAERSK 101W', '2026-04-15', 'MSKU1234567',
         '12500.00', 'Electronics'),
        ('INV-2026-002', 'HBL002', 'MBL-DEF', '18.2', 'Ningbo',
         'Rotterdam', 'COSCO 202E', '2026-04-18', 'TGHU7654321',
         '9800.00', 'Textiles'),
        ('INV-2026-003', 'HBL003', 'MBL-GHI', '32.0', 'Shenzhen',
         'Hamburg', 'ONE 303W', '2026-04-20', 'CMAU9876543',
         '15200.00', 'Machinery'),
    ]

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            for inv_num, hbl, mbl, vol, pol, pod, ves, etd, cnt, gw, desc in invoice_data:
                wb = create_sample_invoice(inv_num, hbl, mbl, vol, pol, pod,
                                           ves, etd, cnt, gw, desc)
                temp_path = Path(tmpdir) / f'{inv_num}.xlsx'
                wb.save(str(temp_path))
                zf.write(str(temp_path), arcname=f'{inv_num}.xlsx')

    print(f'Sample ZIP created: {zip_path}')
    return str(zip_path)


if __name__ == '__main__':
    create_template()
    create_sample_zip()
    print('Sample data created successfully!')
