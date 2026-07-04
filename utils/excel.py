"""Excel utility functions — unified interface for .xls and .xlsx."""

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


_OPENPYXL_FMTS = {".xlsx", ".xlsm", ".xltx"}


def load_workbook_universal(path: str | Path):
    """
    Load any Excel file:
    - .xlsx / .xlsm / .xltx → openpyxl (for template writing)
    - .xls                  → xlrd (for invoice reading)
    Returns (workbook, format_type) where format_type is 'xlsx' or 'xls'.
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext in _OPENPYXL_FMTS:
        wb = load_workbook(str(path), data_only=True)
        return wb, "xlsx"
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        return wb, "xls"
    else:
        raise ValueError(f"不支持的格式: {path.suffix}")


class SheetWrapper:
    """
    Unified wrapper providing the same cell access interface
    for both openpyxl (.xlsx) and xlrd (.xls) sheets.

    Usage:
        wrapper["I29"]       → CellValue with .value
        wrapper.cell(r, c)   → CellValue with .value (1-based row/col)
    """

    def __init__(self, sheet, fmt: str):
        self._sheet = sheet
        self._fmt = fmt

    def __getitem__(self, ref: str):
        col_0, row_0 = _parse_ref(ref)
        return self._cell_at(row_0, col_0)

    def cell(self, row: int, column: int):
        """1-based row/column (openpyxl convention)."""
        return self._cell_at(row - 1, column - 1)

    # ── internal ──

    def _cell_at(self, row_0: int, col_0: int):
        if self._fmt == "xls":
            val = self._sheet.cell_value(row_0, col_0)
        else:
            val = self._sheet.cell(row=row_0 + 1, column=col_0 + 1).value
        return CellValue(val)

    @property
    def name(self):
        if self._fmt == "xls":
            return self._sheet.name
        return self._sheet.title


class CellValue:
    """Thin wrapper so both paths expose a .value attribute."""
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value

    def __repr__(self):
        return f"CellValue({self.value!r})"


def _parse_ref(ref: str) -> tuple[int, int]:
    """Parse 'I29' → (col_0, row_0)."""
    m = re.match(r"([A-Z]+)(\d+)", ref.upper())
    if not m:
        raise ValueError(f"无效单元格引用: {ref}")
    col_0 = column_index_from_string(m.group(1)) - 1
    row_0 = int(m.group(2)) - 1
    return col_0, row_0


def get_active_sheet(wb, fmt: str) -> SheetWrapper:
    """Get the active/default sheet as a SheetWrapper."""
    if fmt == "xls":
        return SheetWrapper(wb.sheet_by_index(0), fmt)
    return SheetWrapper(wb.active, fmt)
