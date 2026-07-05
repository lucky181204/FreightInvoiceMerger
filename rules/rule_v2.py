"""
Rule V2 — Freight Invoice to DAI (6月)

Writes invoice data to a multi-page .xls template.
Each page holds 8 groups of 3 data rows + 1 blank separator.
Data sorted by PO number from Rule1 output's P column ("/" after content).
"""

RULE_META = {
    "id": "rule_v2",
    "name": "Rule2",
    "description": "Freight Invoice to DAI — multi-page .xls output",
    "template_name": "Freight Invoice to DAI (6月）.xls",
    "sort_column": "",
}

from utils.str_parse import after_colon
import re
import os
import time
import traceback
from pathlib import Path

# ── Constants ──
ROWS_PER_PAGE = 47
DATA_START = 10     # Row 10 within each page
DATA_END = 40       # Row 40 within each page
GROUP_ROWS = 3      # 3 rows per group
BLANK_ROWS = 1      # 1 blank row between groups
GROUPS_PER_PAGE = 8  # (40-10+1) / 4 = 7.75 -> 8 groups
MAX_GROUPS_PER_PAGE = 8


def write_sort_data_fn(template_path, output_dir, invoice_dir, sort_file_path,
                        progress_callback=None):
    """Main entry point for Rule2 processing.
    Returns (output_path, success_count, error_count)."""

    def progress(msg):
        if progress_callback:
            progress_callback(msg)
        from utils.logger import logger
        logger.info(msg)

    errors = 0
    success = 0

    # ── Step 1: Read sort file ──
    progress("读取排序文件...")
    sort_order = _read_sort_order(sort_file_path)
    if not sort_order:
        raise ValueError("排序文件中未找到有效PO数据")
    progress(f"排序文件中找到{len(sort_order)}个PO")

    # ── Step 2: Build PO → Invoice file index ──
    progress("建立Invoice索引...")
    invoice_map = _build_invoice_map(invoice_dir)
    progress(f"ZIP中找到{len(invoice_map)}个Invoice")

    # ── Step 3: Sort invoices by sort order ──
    progress("按PO排序...")
    ordered_files = _sort_invoices(sort_order, invoice_map)
    progress(f"匹配到{len(ordered_files)}个Invoice")

    # ── Step 4: Load template ──
    progress("加载模板...")
    import xlrd
    import xlwt
    import xlutils.copy as cp

    # Patch None format_str
    orig_nfr = xlwt.BIFFRecords.NumberFormatRecord.__init__
    def _patched_nfr(self, fmtidx, fmtstr):
        if fmtstr is None:
            fmtstr = ''
        orig_nfr(self, fmtidx, fmtstr)
    xlwt.BIFFRecords.NumberFormatRecord.__init__ = _patched_nfr

    wb = xlrd.open_workbook(template_path, formatting_info=True)
    wb_out = cp.copy(wb)

    # Find the data sheet (last sheet)
    sheet_names = wb.sheet_names()
    data_sheet_name = sheet_names[-1]  # '2026-5' or similar
    ws_out = wb_out.get_sheet(data_sheet_name)
    progress(f"模板Sheet: {data_sheet_name}")

    # ── Step 5: Write data ──
    for idx, file_path in enumerate(ordered_files):
        page_idx = idx // MAX_GROUPS_PER_PAGE
        group_idx = idx % MAX_GROUPS_PER_PAGE

        # Calculate row offset within the template
        # Page start: page_idx * ROWS_PER_PAGE
        # Group start within page: DATA_START + group_idx * (GROUP_ROWS + BLANK_ROWS)
        offset = page_idx * ROWS_PER_PAGE
        base_row = offset + DATA_START + group_idx * (GROUP_ROWS + BLANK_ROWS)

        progress(f"写入第{idx+1}组 (PO: {Path(file_path).stem})")

        try:
            _write_one_group(ws_out, file_path, base_row)
            success += 1
        except Exception as e:
            errors += 1
            progress(f"ERROR: {Path(file_path).name} — {e}")
            continue

    # ── Step 6: Save output ──
    output_name = "Freight Invoice to DAI_Output.xls"
    output_path = Path(output_dir) / output_name
    counter = 1
    while output_path.exists():
        output_path = Path(output_dir) / f"Freight Invoice to DAI_Output({counter}).xls"
        counter += 1

    progress("保存中...")
    wb_out.save(str(output_path))
    progress(f"保存完成: {output_path}")

    return str(output_path), success, errors


def _read_sort_order(sort_file_path: str) -> list[str]:
    """Read PO order from merged.xlsx P column.
    Returns list of PO strings in template order."""
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(sort_file_path, data_only=True)
    ws = wb.active

    po_list = []
    p_col = column_index_from_string('P')

    for r in range(5, ws.max_row + 1):
        p_val = ws.cell(row=r, column=p_col).value
        if p_val:
            po_str = _extract_po(str(p_val))
            if po_str:
                po_list.append(po_str)

    return po_list


def _extract_po(value: str) -> str:
    """Extract PO from 'XXX/PO4500014946-2' → 'PO4500014946-2'"""
    if '/' in value:
        return value.rsplit('/', 1)[-1].strip()
    return value.strip()


def _build_invoice_map(invoice_dir: str) -> dict[str, str]:
    """Build map of PO → invoice file path.
    PO extracted from filename (e.g. ...PO4500014946-2.xls)."""
    invoice_dir = Path(invoice_dir)
    invoice_map = {}

    for f in sorted(invoice_dir.rglob("*.xls")):
        # Extract PO from filename
        m = re.search(r'(PO\d+(?:-\d+)?)', f.stem)
        if m:
            po = m.group(1)
            invoice_map[po] = str(f)

    return invoice_map


def _sort_invoices(sort_order: list[str], invoice_map: dict[str, str]) -> list[str]:
    """Return file paths in sort_order, skipping missing POs."""
    ordered = []
    for po in sort_order:
        if po in invoice_map:
            ordered.append(invoice_map[po])
        else:
            import logging
            logging.getLogger(__name__).warning(f"PO {po} not found in ZIP")
    return ordered


def _write_one_group(ws_out, file_path: str, base_row: int):
    """Write one 3-row group to the template starting at base_row.
    base_row is 1-based Excel row number."""
    import xlrd

    wb = xlrd.open_workbook(file_path)
    # Find the actual data sheet — pick the LAST sheet with >20 rows
    # or the one containing real data (has 'PRODUCT:', 'Shipment of', or similar)
    inv_sheet = None
    for sn in wb.sheet_names():
        s = wb.sheet_by_name(sn)
        # Check for data markers
        has_data = False
        for r in range(min(s.nrows, 15)):
            for c in range(min(s.ncols, 5)):
                cell_val = str(s.cell_value(r, c))
                if any(marker in cell_val for marker in ['PRODUCT:', 'Shipment of', 'Loading Port', 'Container No.']):
                    has_data = True
                    break
        if has_data:
            inv_sheet = s
    # Fallback: use the last sheet
    if inv_sheet is None:
        inv_sheet = wb.sheet_by_index(wb.nsheets - 1)

    # Get invoice data using xlrd 0-indexed access
    def iv(r, c):
        return inv_sheet.cell_value(r - 1, c - 1)  # convert to 1-based

    # ── B10:B12 — B column block (3 rows) ──
    for row_offset in range(3):
        src_row = 10 + row_offset
        tgt_row = base_row + row_offset
        val = iv(src_row, 2)  # column B = col 2
        if val:
            ws_out.write(tgt_row - 1, 1, str(val))

    # ── B6 after colon → C(base_row) ──
    b6 = iv(6, 2)  # B6
    if b6:
        ws_out.write(base_row - 1, 2, after_colon(str(b6)))

    # ── C10:E10 → D(base_row):F(base_row) ──
    for col_offset in range(3):
        src_col = 3 + col_offset  # C=3, D=4, E=5
        tgt_col = 4 + col_offset  # D=4, E=5, F=6
        val = iv(10, src_col)
        if val:
            ws_out.write(base_row - 1, tgt_col - 1, str(val))

    # ── B7 after colon → H(base_row) ──
    b7 = iv(7, 2)  # B7
    if b7:
        ws_out.write(base_row - 1, 7, after_colon(str(b7)))
